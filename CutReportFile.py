# This skript is divide report file with a lot amudim into separate files with one amud in each file.
import ProgramFiles.MyFunctions as MyF
from openpyxl import Workbook, load_workbook


def main():

    rawDataFileName = input("Input original file. It should have *.xlsx extension, and be situated in ResiltCutReport folder.\n")

    inputWB = load_workbook("ResultCutReport\\{}.xlsx".format(rawDataFileName))
    inputWS = inputWB.active

    NewWB = Workbook()
    NewWS = NewWB.active
    currName = inputWS['A1'].value[:4]
    r = 1
    indexString = 0
    for column in inputWS.rows:
        c = 1
        for cell in column:
            if cell.column == 'A' and cell.value and cell.value[:4] != currName:

                NewWB.save("ResultCutReport\\" + currName + ".xlsx")
                NewWB.close()
                indexString = r-1

                currName = cell.value[:4]
                NewWB = Workbook()
                NewWS = NewWB.active

            if cell.column == 'A' and cell.value is None:
                indexString += 1

            NewWS.cell(column=c, row=r - indexString, value=cell.value)
            c += 1
        r += 1

    NewWB.save("ResultCutReport\\" + currName + ".xlsx")
    NewWB.close()
    NewWB.close()


if __name__ == "__main__":
    main()
