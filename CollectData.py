# This skript is assemble data from a lot of files in Input folder
from openpyxl import load_workbook

import ProgramFiles.MyFunctions as MyF

COLUMN_LETTERS = 'CFGHIJKLMNOPQRSTUVWXYZ' # - columns that I'll need

NAME_RB = 'Result.xlsx'

LAST_ROW = 3 # defalute row, where insert result begins

def main():
    dataList = []
    LAST_ROW = 3

    #Iterate through folder Input, defalute value, and treat raw data.
    if MyF.listInputDir():
        for raw_file in MyF.listInputDir():
            rawWB = load_workbook(raw_file)
            sheetList=rawWB.get_sheet_names()
            dataList += MyF.LoadDataFromInput(rawWB[sheetList[0]])
    else:
        print("There are no raw files *.xlsx format in Input folder")
        return 0


    # Treat data in dataList
    MyF.InputLineTreatment(dataList)

    resultWB = MyF.FileResultCreateRewrite(NAME_RB)

    ws = resultWB.active
    ws.title = 'Result Data'

    for row_ind in range(len(dataList)):
        for col_ind in range(len(dataList[row_ind])):
            cur_cell = ws[COLUMN_LETTERS[col_ind] + str(row_ind + LAST_ROW)]
            cur_cell.value = dataList[row_ind][col_ind]


    resultWB.save(NAME_RB)
    rawWB.close()
    resultWB.close()

if __name__ == "__main__":
    main()
