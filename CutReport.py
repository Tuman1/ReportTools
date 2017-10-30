# This script draws up report of cut Excel
from openpyxl import load_workbook

from ProgramFiles import MyFunctions as MyF
from  ProgramFiles.CellSetings import cellStyle

COLUMN_LETTERS = 'CFGHIJKLMNOPQRSTUVWXYZ' # - columns that I'll need

NAME_RB = 'TestResult.xlsx'

LAST_ROW = 3 # defalute row, where insert result begins

def main():
    dataList = []
    LAST_ROW = 3

    MyF.getInputUser();

    #Iterate through folder Input, defalute value, and treat raw data.
    for raw_file in MyF.listInputDir():
        rawWB = load_workbook(raw_file)
        sheetList=rawWB.get_sheet_names()
        dataList = MyF.LoadDataFromInput(rawWB[sheetList[0]])


    # Treat data in dataList
    MyF.InputLineTreatment(dataList)

    resultWB = MyF.FileResultCreateRewrite(NAME_RB)

    ws = resultWB.active
    ws.title = 'Result Data'

    for row_ind in range(len(dataList)):
        for col_ind in range(len(dataList[row_ind])):
            cur_cell = ws[COLUMN_LETTERS[col_ind] + str(row_ind + LAST_ROW)]
            cur_cell.value = dataList[row_ind][col_ind]


    resultWB.save(NAME_RB)
    rawWB.close()
    resultWB.close()

if __name__ == "__main__":
    main()
