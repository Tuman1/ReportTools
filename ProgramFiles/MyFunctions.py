from os import listdir
import string

from openpyxl import Workbook
from openpyxl import load_workbook

LOAD_DATA_REGION_FINISH = 100

SPLIT_VAR = '+'



def StringChanger(arg):  # Поменять имя функции на более что-то конкретное
    temp = arg.split(SPLIT_VAR)
    temp = [x[1:-1] for x in temp]
    return temp


def FileResultCreateRewrite(name_file):
    try:
        temp = load_workbook(name_file)
        return temp
    except FileNotFoundError:
        temp = Workbook()
        temp.save(name_file)
        temp = load_workbook(name_file)
        return temp


def LoadDataFromInput(ActiveSheet, region="A2:F50"):
    """This function is used for load data from input file from region"""
    """Value of default region is for typical file that I'm using at F12"""
    dataList = []
    Rows, Columns = CreateRegionForIteration(region)
    # for column in Columns:
    #     for row in Rows:
    #         CellAdress = column + str(row)
    #         dataList.append(ActiveSheet[CellAdress].value)
    for row_number in range(len(Rows)):
        for column in Columns:
            CellAdress = column + str(Rows[row_number])
            if Columns.index(column) == 0:
                if ActiveSheet[CellAdress].value == None:
                    return dataList
                else:
                    dataList.append([])
            dataList[row_number].append(ActiveSheet[CellAdress].value)

    # for index in range(2, LOAD_DATA_REGION_FINISH):
    #     if ActiveSheet[args[0] + str(index)].value != None:
    #         temp = []
    #         for letter in args:
    #             temp.append(ActiveSheet[letter + str(index)].value)
    #         dataList.append(temp[:])
    #         temp.clear()
    #     else:
    #         break
    return dataList


def InputLineTreatment(arg):
    for row_list in arg:
        split_temp = row_list.pop().split(SPLIT_VAR)
        for EachString in split_temp:
            row_list.append(EachString[1:-1])


def listInputDir(folder='Input'):
    result = [folder + '\\' + x for x in listdir(folder) if x.endswith(".xlsx")]
    if result:
        return result
    else:
        return False

def CreateRegionForIteration(arg):
    """This function creates region for iteration for other functions"""
    temp = arg.split(':')
    FirstLetterRegion = temp[0][0].upper()
    SecondLetterRegion = temp[1][0].upper()
    FirstNumberRegion = int(temp[0][1:])
    SecondNumberRegion = int(temp[1][1:])

    if FirstLetterRegion > SecondLetterRegion or FirstNumberRegion > SecondNumberRegion:
        print('Wrong input, check order of letters and numbers')
        return False

    else:
        column = []
        for letter in string.ascii_uppercase:
            if FirstLetterRegion <= letter <= SecondLetterRegion:
                column.append(letter)
        row = [str(x) for x in range(FirstNumberRegion, SecondNumberRegion+1)]
        return row, column

def getInputUser():
    print("Добрый день")

    decision = input("Хотите продолжить работу по умолчанию? y/n\n")
    while decision !='y' and decision != 'n':
        decision = input("Не подходящий ответ. Выберете y/n\n")

if __name__ == "__main__":
    pass
